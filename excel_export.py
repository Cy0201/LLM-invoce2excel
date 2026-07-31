
# -*- coding: utf-8 -*-
"""excel_export.py —— 通用导出  v7
Sheet1 提取汇总（每条逻辑记录一行，含校验与复核标记）
Sheet2 明细数据（所有表格行摊平，列取并集，带行号）
Sheet3 算术校验（求和/声明/差额，复核状态）
"""
import io
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_THIN = Side(style='thin', color='D9DEE7')
_B = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD = PatternFill('solid', fgColor='2B3A55')
_ALT = PatternFill('solid', fgColor='F1F4F9')
_WARN = PatternFill('solid', fgColor='FCE9CE')
_ERR = PatternFill('solid', fgColor='FBE0DE')
_FIX = PatternFill('solid', fgColor='E4F5EC')

_EVIDENCE_CN = {'native_text': '电子解析', 'vision': 'AI视觉', 'ocr_text': 'OCR文本',
                'failed': '失败'}


def _ev_label(ev):
    parts = [p for p in str(ev or '').split('+') if p]
    return '+'.join(_EVIDENCE_CN.get(p, p) for p in parts) or '—'


def _header(ws, cols, fill=_HEAD):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _B


def _fmt(v, ftype):
    if v is None:
        return ''
    if ftype == 'number':
        try:
            return float(str(v).replace(',', ''))
        except Exception:
            return str(v)
    if ftype == 'checkbox':
        return '✓' if v is True else ('✗' if v is False else str(v))
    if ftype == 'table':
        return '[%d行]' % len(v) if isinstance(v, list) else ''
    return str(v)


def write_excel(records, fields, validations):
    wb = openpyxl.Workbook()
    scalars = [f for f in fields if f['type'] != 'table']
    tables = [f for f in fields if f['type'] == 'table']
    check_fields = [f for f in scalars if f.get('sum_check')]
    any_repaired = any(r.get('_repaired') for r in records)

    # ── Sheet1 汇总 ──────────────────────────────────────────
    ws = wb.active
    ws.title = '提取汇总'
    cols = (['文件名', '页码'] + [f['label'] for f in scalars]
            + [tf['label'] + '(行数)' for tf in tables]
            + ['置信度', '证据来源']
            + (['复核'] if any_repaired else [])
            + ['校验:%s' % f['label'] for f in check_fields] + ['状态'])
    _header(ws, cols)
    for r, rec in enumerate(records, 2):
        is_err = bool(rec.get('_error'))
        checks = [rec.get('_check_%s' % f['key']) for f in check_fields]
        row = ([rec.get('_filename', ''), str(rec.get('_pages', ''))]
               + [_fmt(rec.get(f['key']), f['type']) for f in scalars]
               + [len(rec.get(tf['key']) or []) for tf in tables]
               + [rec.get('_confidence', ''), _ev_label(rec.get('_evidence'))]
               + (['已复核' if rec.get('_repaired') else ''] if any_repaired else [])
               + ['✓' if c is True else ('✗' if c is False else '—') for c in checks]
               + [('❌ ' + str(rec.get('_error', ''))[:60]) if is_err else '✅'])
        any_bad = any(c is False for c in checks)
        fill = (_ERR if is_err else
                (_WARN if any_bad else
                 (_FIX if rec.get('_repaired') else (_ALT if r % 2 == 0 else None))))
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(size=9)
            cell.border = _B
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if fill:
                cell.fill = fill
    for i in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.column_dimensions['A'].width = 24
    ws.freeze_panes = 'C2'

    # ── Sheet2 明细数据 ──────────────────────────────────────
    if tables:
        col_keys = []
        for tf in tables:
            for c in tf['columns']:
                if c['key'] not in col_keys:
                    col_keys.append(c['key'])
        extra = list(OrderedDict.fromkeys(
            k for rec in records for tf in tables
            for row_ in (rec.get(tf['key']) or []) if isinstance(row_, dict)
            for k in row_.keys() if k not in col_keys))
        col_keys += extra
        label_of = {}
        for tf in tables:
            for c in tf['columns']:
                label_of.setdefault(c['key'], '%s(%s)' % (c['label'], c['key']))
        if col_keys:
            ws2 = wb.create_sheet('明细数据')
            _header(ws2, ['文件名', '页码', '表格', '行号']
                    + [label_of.get(k, k) for k in col_keys])
            r2 = 2
            for rec in records:
                for tf in tables:
                    for ridx, row_ in enumerate(rec.get(tf['key']) or [], 1):
                        if not isinstance(row_, dict):
                            continue
                        vals = ([rec.get('_filename', ''), str(rec.get('_pages', '')),
                                 tf['label'], ridx]
                                + [('' if row_.get(k) is None else row_.get(k))
                                   for k in col_keys])
                        for c, v in enumerate(vals, 1):
                            cell = ws2.cell(r2, c, v if isinstance(v, (int, float))
                                            else str(v))
                            cell.font = Font(size=9)
                            cell.border = _B
                            if r2 % 2 == 0:
                                cell.fill = _ALT
                        r2 += 1
            for i in range(1, len(col_keys) + 5):
                ws2.column_dimensions[get_column_letter(i)].width = 15
            ws2.column_dimensions['A'].width = 24
            ws2.freeze_panes = 'E2'

    # ── Sheet3 算术校验 ──────────────────────────────────────
    if validations:
        ws3 = wb.create_sheet('算术校验')
        _header(ws3, ['文件名', '页码', '字段', '表格求和', '声明值', '差额',
                      '结果', '复核'],
                fill=PatternFill('solid', fgColor='7A3E00'))
        for r, v in enumerate(validations, 2):
            vals = [v['file'], v['pages'], v['label'], v['sum_from_table'],
                    v['declared'] if v['declared'] is not None else '',
                    v['diff'] if v['diff'] is not None else '',
                    '✓ 一致' if v['match'] else '✗ 不一致',
                    '已复核' if v.get('repaired') else '']
            for c, val in enumerate(vals, 1):
                cell = ws3.cell(r, c, val)
                cell.font = Font(size=9)
                cell.border = _B
                if not v['match']:
                    cell.fill = _WARN
                elif v.get('repaired'):
                    cell.fill = _FIX
        for i, w in enumerate([24, 10, 14, 14, 14, 10, 12, 10], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
