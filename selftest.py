# -*- coding: utf-8 -*-
"""
selftest.py —— 全链路自测  v7

离线模式（默认，无需网络/网关）：
    python3 selftest.py
  用 reportlab 生成真实 PDF 样本（中文缴款书3页/图片型送货单/空白页），
  以 claude_sim 仿真"称职 AI"跑通 Flask 全链路（SSE→合并→校验→复核→Excel），
  并注入四类故障验证提取阶梯：坏JSON重试 / 截断升额 / 错值自动复核 / 无视觉降级OCR。

在线冒烟（在内网机器上验证真实网关，二选一都行）：
    python3 selftest.py --live            # 文本链路：连通 + 单页真实提取
    python3 selftest.py --live --vision   # 追加视觉链路探测
  网关取环境变量 ANTHROPIC_BASE_URL / ANTHROPIC_AUTH_TOKEN / CLAUDE_MODEL；
  想对照 Claude：把三者指向 https://api.anthropic.com 与 Claude 模型即可。
"""
import io
import re
import sys
import json
import time
import argparse
import traceback

PASS, FAIL = 0, []


def check(name, cond, detail=''):
    global PASS
    if cond:
        PASS += 1
        print('  ✓ %s' % name)
    else:
        FAIL.append(name)
        print('  ✗ %s  %s' % (name, detail))


# ══════════════════════════════════════════════════════════════
#  样本构造（真实 PDF）
# ══════════════════════════════════════════════════════════════
def build_tax_pdf():
    """3页中文缴款书：p1=单据A(2行)；p2=单据B首页(3行,无合计)；p3=B续页(2行+合计)。"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    W, H = A4
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    XS = [40, 70, 150, 285, 335, 380, 470, 515]
    HEADS = ['序号', '税号', '货物名称', '数量', '单位', '完税价格', '税率', '税款金额']

    def head(no, decl, date):
        c.setFont('STSong-Light', 14)
        c.drawString(150, H - 60, '中华人民共和国海关进口关税专用缴款书')
        c.setFont('STSong-Light', 9)
        c.drawString(40, H - 88, '关别：深圳海关')
        c.drawString(200, H - 88, '填发日期：%s' % date)
        c.drawString(400, H - 88, '号码：%s' % no)
        c.drawString(40, H - 106, '缴款单位：深圳市朗华供应链服务有限公司')
        c.drawString(340, H - 106, '报关单编号：%s' % decl)

    def table(rows, y0, with_head=True):
        y = y0
        c.setFont('STSong-Light', 8.5)
        if with_head:
            for x, t in zip(XS, HEADS):
                c.drawString(x, y, t)
            y -= 16
        for r in rows:
            for x, v in zip(XS, r):
                c.drawString(x, y, v)
            y -= 15
        return y

    # p1 单据A
    head('H0001', '531620261034567890', '2026年07月12日')
    y = table([('1', '84713000', '便携式电脑', '120', '台', '960,000.00', '10', '96,000.00'),
               ('2', '85285910', '液晶显示器', '300', '台', '450,000.00', '8', '36,000.00')],
              H - 132)
    c.setFont('STSong-Light', 9)
    c.drawString(40, y - 8, '合计（¥）：132,000.00')
    c.showPage()

    # p2 单据B 首页（无合计）
    head('H0002', '531620269876543210', '2026年07月13日')
    y = table([('1', '39269090', '塑料支架', '5,000', '个', '25,000.00', '10', '1,000.00'),
               ('2', '85044090', '电源模块', '800', '个', '40,010.00', '5', '2,000.50'),
               ('3', '84733090', '主板配件', '1,200', '个', '60,000.00', '5', '3,000.00')],
              H - 132)
    c.setFont('STSong-Light', 8)
    c.drawString(430, 40, '（接下页）')
    c.showPage()

    # p3 单据B 续页（无抬头编号，有合计）
    c.setFont('STSong-Light', 10)
    c.drawString(40, H - 60, '（续页）货物明细（接上页）')
    y = table([('4', '73182200', '垫圈', '10,000', '个', '5,002.50', '10', '500.25'),
               ('5', '40169390', '密封圈', '9,995', '个', '14,992.50', '10', '1,499.25')],
              H - 90, with_head=True)
    c.setFont('STSong-Light', 9)
    c.drawString(40, y - 8, '合计（¥）：8,000.00')
    c.showPage()
    c.save()
    return buf.getvalue()


DELIVERY_OCR = (
    '送货单\n'
    '送货单号：DN-2026-0715    送货日期：2026-07-15\n'
    '供应商：ACME贸易有限公司\n'
    '收货单位：珠海富创电子厂\n'
    '序号   品名   规格   单位   数量   备注\n'
    '1   USB数据线   1.5m   条   10   无\n'
    '2   电源适配器   65W   个   25   加急\n'
    '总数量：35')


def build_scan_pdf():
    """图片型 PDF（无文本层）。内容用色块模拟——OCR 由测试桩提供 DELIVERY_OCR。"""
    from PIL import Image, ImageDraw
    img = Image.new('RGB', (1240, 1754), 'white')
    d = ImageDraw.Draw(img)
    for i, y in enumerate(range(120, 1400, 90)):
        d.rectangle([100, y, 1100, y + 34], outline='black', width=3)
        d.line([100, y + 17, 400 + (i % 4) * 150, y + 17], fill='black', width=5)
    buf = io.BytesIO()
    img.save(buf, format='PDF')
    return buf.getvalue()


def build_blank_pdf():
    from PIL import Image
    buf = io.BytesIO()
    Image.new('RGB', (1240, 1754), 'white').save(buf, format='PDF')
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════
#  SSE 工具
# ══════════════════════════════════════════════════════════════
def sse_events(raw):
    out = []
    for chunk in raw.decode('utf-8').split('\n\n'):
        chunk = chunk.strip()
        if chunk.startswith('data: '):
            out.append(json.loads(chunk[6:]))
    return out


# ══════════════════════════════════════════════════════════════
#  离线全链路
# ══════════════════════════════════════════════════════════════
def offline():
    import app as APP
    import pdf_utils as P
    import presets as PS
    from merge import (normalize_fields, parse_columns, merge_pages, to_num)
    from extractor import coerce, norm_date
    from claude_sim import SimAI, CommonSimAI
    import common_mode as CM

    print('\n[单元] 基础构件')
    cols = parse_columns('列：seq 序号, hs_code 税号, name 货物名称')
    check('列声明解析', [c['key'] for c in cols] == ['seq', 'hs_code', 'name'], cols)
    check('中文列名兜底', parse_columns('列：金额, tax 税额')[0]['key'] == 'col1')
    check('日期归一', norm_date('２０２６年7月3日') == '2026-07-03', norm_date('２０２６年7月3日'))
    check('数值归一', coerce('￥１２,３45.60元', 'number') == 12345.6)
    check('中文文件名保留', APP.safe_name('目录/税票 A.pdf') == '税票 A.pdf')
    check('重名去重', APP._uniq_names(['a.pdf', 'a.pdf'])[1] == 'a(1).pdf')

    from ai_client import GatewayConfig, AIClient, AIResponseError
    qwen_client = AIClient(GatewayConfig('http://example.invalid', 'x', 'qwen-test'))
    check('no_think追加在Qwen用户消息末尾',
          qwen_client._user('请输出JSON').endswith('/no_think'))
    retry_calls = []
    def thinking_once(kind, system, user, max_tokens, image_b64=None):
        retry_calls.append((user, qwen_client._prefill_ok))
        if len(retry_calls) == 1:
            raise AIResponseError('只有思考', only_thinking=True)
        return '{"ok":true}', 'end_turn'
    qwen_client._raw_call = thinking_once
    text, _ = qwen_client.call('text', 'system', 'user', 100)
    check('仅思考自动关闭预填并重试',
          text == '{"ok":true}' and len(retry_calls) == 2 and
          retry_calls[1][1] is False and retry_calls[1][0].endswith('/no_think'),
          retry_calls)

    common_fields = CM.normalize_common_fields([
        {'key': 'common_value', 'label': '统一字段', 'type': 'text',
         'description': '不同文档中的同一业务含义'}])
    mixed_pages = [
        {'_filename': 'mix.pdf', '_page': 1, '_document_type': '类型甲',
         '_document_no': 'A-1', '_is_continuation': False,
         'common_value': '甲值', '_field_meta': {'common_value': {'status': 'found'}},
         '_confidence': 'high', '_evidence': 'native_text'},
        {'_filename': 'mix.pdf', '_page': 2, '_document_type': '类型乙',
         '_document_no': 'B-1', '_is_continuation': False,
         'common_value': '乙值', '_field_meta': {'common_value': {'status': 'found'}},
         '_confidence': 'high', '_evidence': 'native_text'},
        {'_filename': 'mix.pdf', '_page': 3, '_document_type': '类型甲',
         '_document_no': None, '_is_continuation': True,
         'common_value': '甲值', '_field_meta': {'common_value': {'status': 'found'}},
         '_confidence': 'high', '_evidence': 'native_text'},
    ]
    common_recs, common_issues = CM.assemble_common_pages(mixed_pages, common_fields)
    check('异构交错页按同类型续页归组',
          len(common_recs) == 2 and common_recs[0]['_pages'] == '1+3', common_recs)
    check('通用归组无票据类型规则', not common_issues, common_issues)

    detail_fields = CM.normalize_common_fields([
        {'key': 'transactions', 'label': '交易明细', 'type': 'table',
         'description': '不同版式对账单中的逐笔交易',
         'columns': [
             {'key': 'trade_date', 'label': '交易日期', 'type': 'date',
              'description': '该笔交易入账日期'},
             {'key': 'amount', 'label': '发生额', 'type': 'number',
              'description': '该笔交易金额'},
         ]}])
    detail_pages = [
        {'_filename': 'bank.pdf', '_page': 1, '_segment_anchor': 1,
         '_segment_confidence': 'high', '_document_type': '银行对账单',
         'transactions': [{'trade_date': '2026-01-01', 'amount': 10}],
         '_field_meta': {'transactions': {'status': 'found', 'confidence': 'high'}},
         '_confidence': 'high', '_evidence': 'native_text'},
        {'_filename': 'bank.pdf', '_page': 2, '_segment_anchor': 1,
         '_segment_confidence': 'high', '_document_type': '银行对账单',
         'transactions': [{'trade_date': '2026-01-02', 'amount': -3}],
         '_field_meta': {'transactions': {'status': 'found', 'confidence': 'high'}},
         '_confidence': 'high', '_evidence': 'native_text'},
    ]
    detail_recs, detail_issues = CM.assemble_common_pages(detail_pages, detail_fields)
    check('异构明细表跨页统一归并',
          len(detail_recs) == 1 and len(detail_recs[0]['transactions']) == 2,
          detail_recs)
    from common_export import write_common_excel
    import openpyxl
    detail_book = openpyxl.load_workbook(io.BytesIO(
        write_common_excel(detail_recs, detail_fields, detail_issues)))
    check('统一明细表生成独立Excel工作表',
          any(name.startswith('明细-交易明细') for name in detail_book.sheetnames),
          detail_book.sheetnames)

    long_pages = []
    for page_no in range(1, 86):
        is_same = page_no in (1, 85)
        long_pages.append({
            '_filename': 'long-mix.pdf', '_page': page_no,
            '_document_type': '类型甲' if is_same else '类型%02d' % page_no,
            '_document_no': 'LONG-A' if is_same else None,
            '_page_role': 'first' if page_no == 1 else
                          ('continuation' if page_no == 85 else 'single'),
            '_is_continuation': page_no == 85,
            '_identity_hints': [], '_page_summary': '',
            'common_value': '甲值' if is_same else '值%02d' % page_no,
            '_field_meta': {'common_value': {'status': 'found', 'confidence': 'high'}},
            '_confidence': 'high', '_evidence': 'native_text',
        })

    def splitting_boundary_ai(kind, system, user, max_tokens, image_b64=None):
        payload = json.loads(user)
        return json.dumps({'pages': [
            {'page': p['page'], 'anchor_page': p['page'],
             'document_type': p['initial_document_type'],
             'document_no': p['initial_document_no'],
             'confidence': 'high', 'reason': '故意逐页拆分'}
            for p in payload['pages']
        ]}, ensure_ascii=False), 'end_turn'

    CM.refine_document_boundaries(splitting_boundary_ai, long_pages, common_fields)
    long_recs, _ = CM.assemble_common_pages(long_pages, common_fields)
    check('长文件跨分段同编号页面仍归为一份',
          any(r['_pages'] == '1+85' for r in long_recs),
          [r['_pages'] for r in long_recs])

    fields_nc = normalize_fields([
        {'key': 'title', 'label': '标题', 'type': 'text'},
        {'key': 't', 'label': '明细', 'type': 'table', 'description': '列：a 甲'}])
    pages = [{'_filename': 'x.pdf', '_page': 1, 'title': 'T', 't': [{'a': 1}],
              '_is_continuation': False, '_evidence': 'native_text'},
             {'_filename': 'x.pdf', '_page': 2, 'title': None, 't': [{'甲': 2}],
              '_is_continuation': True, '_evidence': 'native_text'}]
    recs, _ = merge_pages(pages, fields_nc)
    check('无编号字段也能靠续页链合并', len(recs) == 1 and recs[0]['_pages'] == '1+2')
    check('行键中文→key重映射', recs[0]['t'][1].get('a') == 2, recs[0]['t'])

    print('\n[样本] 生成真实 PDF 并验证三态判定')
    tax_pdf = build_tax_pdf()
    scan_pdf = build_scan_pdf()
    blank_pdf = build_blank_pdf()
    pf = P.PdfFile('税票样例.pdf', tax_pdf).scan()
    check('电子件3页判定 native', [pg.kind for pg in pf.pages] == ['native'] * 3,
          [pg.kind for pg in pf.pages])
    check('版式文本含列对齐(≥2空格)', '   ' in pf.pages[0].text)
    pf.close()
    check('图片型判定 scanned',
          P.PdfFile('s.pdf', scan_pdf).scan().pages[0].kind == 'scanned')
    check('空白页判定 blank',
          P.PdfFile('b.pdf', blank_pdf).scan().pages[0].kind == 'blank')
    from PIL import Image, ImageDraw
    image_buf = io.BytesIO()
    image = Image.new('RGB', (900, 1200), 'white')
    draw = ImageDraw.Draw(image)
    draw.rectangle((80, 80, 820, 1100), outline='black', width=8)
    draw.text((120, 140), 'IMAGE RECEIPT 2026-07-12 TOTAL 100', fill='black')
    image.save(image_buf, format='PNG')
    image_bytes = image_buf.getvalue()
    image_file = P.open_file('票据图片.png', image_bytes).scan()
    check('普通图片按扫描页接入', image_file.page_count == 1 and
          image_file.pages[0].kind == 'scanned' and image_file.render(0).size[0] > 0)

    # OCR 测试桩：扫描页由桩提供文本（模拟 rapidocr）
    P.ocr_layout = lambda img, limit=6000: DELIVERY_OCR

    client = APP.app.test_client()

    print('\n[链路1] 缴款书3页+空白：坏JSON重试 / 截断升额 / 错值自动复核 / 续页合并')
    sim = SimAI(pages_bad_json={('税票样例.pdf', 1)},
                pages_truncate={('税票样例.pdf', 2)},
                wrong_scalar={('税票样例.pdf', 3): ('total_amount', 100.0)})
    APP._AI_CALL_OVERRIDE = sim
    fields = PS.PRESETS['报关税票/缴款书']
    resp = client.post('/extract', data={
        'fields': json.dumps(fields, ensure_ascii=False),
        'files': [(io.BytesIO(tax_pdf), '税票样例.pdf'),
                  (io.BytesIO(blank_pdf), '空白.pdf')]})
    evs = sse_events(resp.data)
    start = next(e for e in evs if e['type'] == 'start')
    res = next(e for e in evs if e['type'] == 'result')
    check('总页数=4 空白=1', start['total_pages'] == 4 and start['blank_pages'] == 1, start)
    check('空白页被跳过(page事件mode=blank)',
          any(e.get('mode') == 'blank' for e in evs if e['type'] == 'page'))
    check('逻辑记录=2（A单页 + B跨页）', res['records'] == 2, res['records'])
    rows = res['rows']
    rec_b = next(r for r in rows if r['_pages'] == '2+3')
    rec_a = next(r for r in rows if r['_pages'] == '1')
    check('B记录跨页合并 2+3', rec_b['declaration_no'] == '531620269876543210', rec_b)
    check('B明细共5行(3+2)', len(rec_b['_tables']['goods_table']) == 5)
    check('B合计取末页(carry last)=8000', rec_b['total_amount'] == 8000.0,
          rec_b['total_amount'])
    check('A合计=132000', rec_a['total_amount'] == 132000.0)
    check('校验2/2全过（复核后）',
          res['validation']['passed'] == 2 and not res['validation']['issues'],
          res['validation'])
    check('B记录带复核标记', rec_b['_repaired'] is True)
    check('A记录未触发复核', rec_a['_repaired'] is False)
    p1_extract_calls = [c for c in sim.calls if c[1] == ('税票样例.pdf', 1)]
    check('坏JSON触发提醒重试(p1≥2次)', len(p1_extract_calls) >= 2, sim.calls)
    check('截断触发升额到3276800(p2)',
          any(c[1] == ('税票样例.pdf', 2) and c[2] == 3276800 for c in sim.calls))
    check('评级聚合', rec_b['_confidence'] == 'high')
    dl = client.get('/download?job=' + res['job'])
    check('Excel按job下载', dl.status_code == 200 and dl.data[:2] == b'PK',
          dl.status_code)

    print('\n[链路2] 图片型送货单 · 无视觉网关 → 自动降级OCR文本')
    APP._AI_CALL_OVERRIDE = SimAI(no_vision=True)
    resp = client.post('/extract', data={
        'fields': json.dumps(PS.PRESETS['送货单'], ensure_ascii=False),
        'files': [(io.BytesIO(scan_pdf), '送货单扫描.pdf')]})
    evs = sse_events(resp.data)
    res = next(e for e in evs if e['type'] == 'result')
    row = res['rows'][0]
    check('降级为OCR文本证据', row['_evidence'] == 'ocr_text', row['_evidence'])
    check('单号提取', row['delivery_no'] == 'DN-2026-0715', row)
    check('数量求和校验通过 35',
          res['validation']['passed'] == 1 and row['total_qty'] == 35.0,
          res['validation'])

    print('\n[链路3] 分析与探测端点')
    APP._AI_CALL_OVERRIDE = SimAI()
    r = client.post('/api/analyze', data={'file': (io.BytesIO(tax_pdf), '税票样例.pdf')})
    d = r.get_json()
    check('分析返回字段方案(非回退)',
          d.get('doc_type') == '报关税票/缴款书' and 'warning' not in d
          and len(d.get('fields', [])) >= 5,
          {k: d.get(k) for k in ('doc_type', 'warning')})
    check('分析报告页面构成', d['pages'] == {'total': 3, 'native': 3,
                                       'scanned': 0, 'blank': 0}, d.get('pages'))
    r = client.post('/api/ping', data={})
    d = r.get_json()
    check('ping 文本/视觉探测', d['text_ok'] and d['vision_ok'], d)

    print('\n[链路4] 异构票据 · AI归纳公共字段 → 统一提取 → 非连续续页归组')
    APP._AI_CALL_OVERRIDE = CommonSimAI()
    r = client.post('/api/common/analyze', data={
        'files': [(io.BytesIO(tax_pdf), '混合文档.pdf')]})
    d = r.get_json()
    check('AI公共字段分析返回方案', r.status_code == 200 and len(d.get('fields', [])) == 2,
          d)
    check('公共字段保留语义来源与覆盖率',
          d['fields'][0].get('source_variants') == ['日期栏'] and
          d['fields'][0].get('coverage') == 1.0, d['fields'][0])
    resp = client.post('/common/extract', data={
        'fields': json.dumps(d['fields'], ensure_ascii=False),
        'document_types': json.dumps(d['document_types'], ensure_ascii=False),
        'files': [(io.BytesIO(tax_pdf), '混合文档.pdf')]})
    evs = sse_events(resp.data)
    res = next((e for e in evs if e['type'] == 'result'), None)
    check('统一提取端点完成', res is not None and res.get('mode') == 'common', res)
    check('统一模式续页归组为两份逻辑文档',
          res and res['records'] == 2 and any(r['_pages'] == '2+3' for r in res['rows']),
          res and res['rows'])
    check('统一结果保留原字段与证据',
          res and res['rows'][0]['_field_meta']['document_date']['source_label'] == '日期栏')
    dl = client.get('/download?job=' + res['job']) if res else None
    check('统一模式Excel下载', dl is not None and dl.status_code == 200 and dl.data[:2] == b'PK')

    APP._AI_CALL_OVERRIDE = CommonSimAI()
    resp = client.post('/common/run', data={
        'files': [(io.BytesIO(tax_pdf), '一键混合文档.pdf')]})
    evs = sse_events(resp.data)
    schema = next((e for e in evs if e['type'] == 'schema'), None)
    one_click = next((e for e in evs if e['type'] == 'result'), None)
    check('一键模式自动归纳字段并继续提取',
          schema is not None and len(schema.get('fields', [])) == 2 and
          one_click is not None and one_click.get('records') == 2,
          {'schema': schema, 'result': one_click})

    APP._AI_CALL_OVERRIDE = CommonSimAI()
    resp = client.post('/common/extract', data={
        'fields': json.dumps(d['fields'], ensure_ascii=False),
        'document_types': json.dumps(d['document_types'], ensure_ascii=False),
        'files': [(io.BytesIO(image_bytes), '直接上传票据.png')]})
    image_result = next((e for e in sse_events(resp.data) if e['type'] == 'result'), None)
    check('图片直接上传统一提取全链路',
          image_result is not None and image_result.get('records') == 1,
          image_result)

    print('\n[边界] 破损文件与空字段')
    APP._AI_CALL_OVERRIDE = SimAI()
    resp = client.post('/extract', data={
        'fields': json.dumps(PS.PRESETS['送货单'], ensure_ascii=False),
        'files': [(io.BytesIO(b'not a pdf'), '坏.pdf'),
                  (io.BytesIO(scan_pdf), '好.pdf')]})
    evs = sse_events(resp.data)
    res = next((e for e in evs if e['type'] == 'result'), None)
    check('坏文件不拖垮同批好文件', res is not None and res['records'] == 1,
          [e for e in evs if e['type'] != 'page'])
    check('坏文件有错误事件',
          any(e.get('mode') == 'error' and e.get('filename') == '坏.pdf'
              for e in evs if e['type'] == 'page'))
    r = client.post('/extract', data={'fields': '[]',
                                      'files': [(io.BytesIO(scan_pdf), 'x.pdf')]})
    check('空字段定义→400', r.status_code == 400)
    r = client.get('/')
    check('首页模板渲染', r.status_code == 200
           and '票据智能提取台' in r.get_data(as_text=True)
           and '求和校验' in r.get_data(as_text=True)
           and '不同票据统一字段' in r.get_data(as_text=True))

    APP._AI_CALL_OVERRIDE = None


# ══════════════════════════════════════════════════════════════
#  在线冒烟（内网真实网关 / Claude 对照）
# ══════════════════════════════════════════════════════════════
def live(vision):
    import os
    import pdf_utils as P
    import presets as PS
    from ai_client import GatewayConfig, AIClient
    from merge import normalize_fields, merge_pages
    from extractor import extract_page

    cfg = GatewayConfig(os.environ.get('ANTHROPIC_BASE_URL',
                                       'http://172.16.92.211:23323'),
                        os.environ.get('ANTHROPIC_AUTH_TOKEN', ''),
                        os.environ.get('CLAUDE_MODEL', 'qwen3.6'))
    print('网关 %s · 模型 %s' % (cfg.base_url, cfg.model))
    ai = AIClient(cfg).call
    t0 = time.time()
    txt, _ = ai('text', '你是回显器。', '只回复两个字：OK', 2400)
    print('  ✓ 文本连通 %dms → %r' % ((time.time() - t0) * 1000, txt[:20]))

    tax = build_tax_pdf()
    pf = P.PdfFile('冒烟样例.pdf', tax).scan()
    fields = normalize_fields(PS.PRESETS['报关税票/缴款书'])
    print('  · 用真实模型提取样例第1页（电子件文本路线）…')
    t0 = time.time()
    page = extract_page(fields, ai, kind='native', fname='冒烟样例.pdf',
                        page_no=1, total=3, text=pf.pages[0].text, ctx='[冒烟] ')
    dt = time.time() - t0
    page.update(_filename='冒烟样例.pdf', _page=1)
    recs, vals = merge_pages([page], fields)
    print(json.dumps(recs[0], ensure_ascii=False, indent=2, default=str)[:1200])
    ok = vals and vals[0]['match']
    print('  %s 单页提取 %.1fs · 求和校验 %s（声明 %s / 求和 %s）'
          % ('✓' if ok else '✗', dt, '一致' if ok else '不一致',
             vals[0]['declared'] if vals else '-',
             vals[0]['sum_from_table'] if vals else '-'))
    if vision:
        print('  · 视觉链路探测（首页渲染成图后提问）…')
        b64 = P.pil_to_b64(pf.render(0, target_side=1400), quality=85)
        t0 = time.time()
        txt, _ = ai('vision', '你是文档助手。',
                    '这页是什么单据？只回答单据名称。', 6400, image_b64=b64)
        print('  ✓ 视觉连通 %.1fs → %r' % (time.time() - t0, txt[:60]))
    pf.close()


# ══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--live', action='store_true', help='连真实网关冒烟')
    ap.add_argument('--vision', action='store_true', help='冒烟时追加视觉探测')
    a = ap.parse_args()
    try:
        if a.live:
            live(a.vision)
        else:
            offline()
    except Exception:
        traceback.print_exc()
        FAIL.append('未捕获异常')
    print('\n' + '=' * 46)
    if a.live:
        print('在线冒烟结束（结果见上）')
    elif FAIL:
        print('✗ 通过 %d 项，失败 %d 项：%s' % (PASS, len(FAIL), '；'.join(FAIL)))
        sys.exit(1)
    else:
        print('✓ 全部 %d 项自测通过' % PASS)
