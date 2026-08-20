# -*- coding: utf-8 -*-
"""异构票据统一字段模式的 Excel 导出。"""
import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEAD = PatternFill('solid', fgColor='2B3A55')
_ALT = PatternFill('solid', fgColor='FAF7EF')
_WARN = PatternFill('solid', fgColor='FCEFD9')


def _header(ws, labels):
    for col, label in enumerate(labels, 1):
        cell = ws.cell(1, col, label)
        cell.font = Font(color='FFFFFF', bold=True, size=10)
        cell.fill = _HEAD
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def write_common_excel(records, fields, issues):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '统一汇总'
    labels = ['文档ID', '来源文件', '页码', 'AI识别类型', '逻辑文档编号'] + \
             [f['label'] for f in fields] + ['提取置信度', '边界置信度', '状态']
    _header(ws, labels)
    for r, rec in enumerate(records, 2):
        meta = rec.get('_field_meta') or {}
        ambiguous = any(x.get('status') == 'ambiguous' for x in meta.values())
        vals = [rec.get('_document_id', ''), rec.get('_filename', ''),
                str(rec.get('_pages', '')), rec.get('_document_type', ''),
                rec.get('_document_no') or '']
        vals += [('%d行' % len(rec.get(f['key']) or [])) if f['type'] == 'table'
                 else (rec.get(f['key']) if rec.get(f['key']) is not None else '')
                 for f in fields]
        boundary_low = rec.get('_segmentation_confidence') == 'low'
        vals += [rec.get('_confidence', ''), rec.get('_segmentation_confidence', ''),
                 '失败' if rec.get('_error') else
                 ('待复核' if ambiguous or boundary_low else '完成')]
        for c, value in enumerate(vals, 1):
            cell = ws.cell(r, c, value if isinstance(value, (int, float)) else str(value))
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if ambiguous:
                cell.fill = _WARN
            elif r % 2 == 0:
                cell.fill = _ALT
    widths = [14, 26, 12, 18, 22] + [18] * len(fields) + [12, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws2 = wb.create_sheet('字段证据')
    _header(ws2, ['文档ID', '来源文件', '页码', 'AI识别类型', '统一字段', '提取值',
                  '原字段名', '证据原文', '字段置信度', '状态', '证据页'])
    r2 = 2
    for rec in records:
        meta = rec.get('_field_meta') or {}
        for f in fields:
            m = meta.get(f['key']) or {}
            vals = [rec.get('_document_id', ''), rec.get('_filename', ''),
                    str(rec.get('_pages', '')), rec.get('_document_type', ''), f['label'],
                    ('%d行' % len(rec.get(f['key']) or [])) if f['type'] == 'table'
                    else (rec.get(f['key']) if rec.get(f['key']) is not None else ''),
                    m.get('source_label') or '', m.get('evidence') or '',
                    m.get('confidence') or '', m.get('status') or '',
                    m.get('source_page') or '']
            for c, value in enumerate(vals, 1):
                cell = ws2.cell(r2, c, value if isinstance(value, (int, float)) else str(value))
                cell.font = Font(size=9)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if m.get('status') == 'ambiguous':
                    cell.fill = _WARN
                elif r2 % 2 == 0:
                    cell.fill = _ALT
            r2 += 1
    for i, width in enumerate([14, 26, 12, 18, 18, 20, 18, 42, 12, 15, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    # 每个统一明细表单独展开，便于不同银行流水/商品明细直接做筛选和透视。
    used_titles = set(wb.sheetnames)
    for field in (f for f in fields if f.get('type') == 'table'):
        title = _sheet_title('明细-' + field['label'], used_titles)
        used_titles.add(title)
        wt = wb.create_sheet(title)
        columns = field.get('columns') or []
        _header(wt, ['文档ID', '来源文件', '页码', 'AI识别类型'] +
                [c['label'] for c in columns])
        rr = 2
        for rec in records:
            for row in rec.get(field['key']) or []:
                vals = [rec.get('_document_id', ''), rec.get('_filename', ''),
                        str(rec.get('_pages', '')), rec.get('_document_type', '')]
                vals += [row.get(c['key'], '') for c in columns]
                for cc, value in enumerate(vals, 1):
                    cell = wt.cell(rr, cc, value if isinstance(value, (int, float)) else str(value))
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    if rr % 2 == 0:
                        cell.fill = _ALT
                rr += 1
        for i, width in enumerate([14, 26, 12, 18] + [18] * len(columns), 1):
            wt.column_dimensions[get_column_letter(i)].width = width

    wbnd = wb.create_sheet('文档边界')
    _header(wbnd, ['文档ID', '来源文件', '归组页码', 'AI识别类型', '文档编号',
                   '边界置信度', '判断依据'])
    for rr, rec in enumerate(records, 2):
        vals = [rec.get('_document_id', ''), rec.get('_filename', ''),
                str(rec.get('_pages', '')), rec.get('_document_type', ''),
                rec.get('_document_no') or '', rec.get('_segmentation_confidence', ''),
                rec.get('_segmentation_reason', '')]
        for cc, value in enumerate(vals, 1):
            cell = wbnd.cell(rr, cc, str(value))
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if rec.get('_segmentation_confidence') == 'low':
                cell.fill = _WARN
            elif rr % 2 == 0:
                cell.fill = _ALT
    for i, width in enumerate([14, 26, 14, 18, 22, 12, 60], 1):
        wbnd.column_dimensions[get_column_letter(i)].width = width

    if issues:
        ws3 = wb.create_sheet('异常记录')
        _header(ws3, ['文档ID', '字段', '说明'])
        for r, issue in enumerate(issues, 2):
            for c, value in enumerate([issue.get('document_id', ''), issue.get('field', ''),
                                       issue.get('message', '')], 1):
                cell = ws3.cell(r, c, str(value))
                cell.font = Font(size=9)
                cell.fill = _WARN
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws3.column_dimensions['A'].width = 16
        ws3.column_dimensions['B'].width = 18
        ws3.column_dimensions['C'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_mixed_excel(records, schemas, issues):
    """混合异构模式导出：索引、按类型汇总、按类型明细和异常。"""
    wb = openpyxl.Workbook()
    index = wb.active
    index.title = '文档索引'
    _header(index, ['文档ID', '来源文件', '页码', '文档类型', '文档编号',
                    '字段方案', '提取置信度', '边界置信度', '状态'])
    for r, rec in enumerate(records, 2):
        meta = rec.get('_field_meta') or {}
        ambiguous = any(m.get('status') == 'ambiguous' for m in meta.values())
        boundary_low = rec.get('_segmentation_confidence') == 'low'
        vals = [rec.get('_document_id', ''), rec.get('_filename', ''),
                str(rec.get('_pages', '')), rec.get('_schema_type') or rec.get('_document_type', ''),
                rec.get('_document_no') or '', rec.get('_schema_type', ''),
                rec.get('_confidence', ''), rec.get('_segmentation_confidence', ''),
                '失败' if rec.get('_error') else
                ('待复核' if ambiguous or boundary_low else '完成')]
        for c, value in enumerate(vals, 1):
            cell = index.cell(r, c, str(value))
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if ambiguous or boundary_low:
                cell.fill = _WARN
            elif r % 2 == 0:
                cell.fill = _ALT
    for i, width in enumerate([16, 26, 14, 20, 22, 20, 12, 12, 12], 1):
        index.column_dimensions[get_column_letter(i)].width = width

    used = set(wb.sheetnames)
    for schema in schemas:
        dtype = schema.get('document_type') or '未知'
        fields = schema.get('fields') or []
        type_records = [r for r in records if
                        str(r.get('_schema_type') or r.get('_document_type')) == str(dtype)]
        title = _sheet_title('汇总-' + dtype, used)
        used.add(title)
        ws = wb.create_sheet(title)
        _header(ws, ['文档ID', '来源文件', '页码', '文档编号'] +
                [f['label'] for f in fields] + ['置信度', '边界置信度', '状态'])
        for rr, rec in enumerate(type_records, 2):
            meta = rec.get('_field_meta') or {}
            ambiguous = any(m.get('status') == 'ambiguous' for m in meta.values())
            vals = [rec.get('_document_id', ''), rec.get('_filename', ''),
                    str(rec.get('_pages', '')), rec.get('_document_no') or '']
            vals += [('%d行' % len(rec.get(f['key']) or [])) if f['type'] == 'table'
                     else (rec.get(f['key']) if rec.get(f['key']) is not None else '')
                     for f in fields]
            vals += [rec.get('_confidence', ''), rec.get('_segmentation_confidence', ''),
                     '失败' if rec.get('_error') else ('待复核' if ambiguous else '完成')]
            for c, value in enumerate(vals, 1):
                cell = ws.cell(rr, c, str(value))
                cell.font = Font(size=9)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if ambiguous or rec.get('_segmentation_confidence') == 'low':
                    cell.fill = _WARN
                elif rr % 2 == 0:
                    cell.fill = _ALT
        for i, width in enumerate([16, 26, 14, 22] + [18] * len(fields) + [12, 12, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        for field in (f for f in fields if f.get('type') == 'table'):
            dtitle = _sheet_title('%s-%s' % (dtype, field['label']), used)
            used.add(dtitle)
            detail = wb.create_sheet(dtitle)
            columns = field.get('columns') or []
            _header(detail, ['文档ID', '来源文件', '页码', '文档编号'] +
                    [c['label'] for c in columns])
            rr = 2
            for rec in type_records:
                for row in rec.get(field['key']) or []:
                    vals = [rec.get('_document_id', ''), rec.get('_filename', ''),
                            str(rec.get('_pages', '')), rec.get('_document_no') or '']
                    vals += [row.get(c['key'], '') for c in columns]
                    for c, value in enumerate(vals, 1):
                        cell = detail.cell(rr, c, str(value))
                        cell.font = Font(size=9)
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                    rr += 1

    boundary = wb.create_sheet('文档边界')
    _header(boundary, ['文档ID', '来源文件', '页码', '文档类型', '边界置信度', '判断依据'])
    for rr, rec in enumerate(records, 2):
        vals = [rec.get('_document_id', ''), rec.get('_filename', ''),
                str(rec.get('_pages', '')), rec.get('_schema_type') or '',
                rec.get('_segmentation_confidence', ''), rec.get('_segmentation_reason', '')]
        for c, value in enumerate(vals, 1):
            cell = boundary.cell(rr, c, str(value))
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if rec.get('_segmentation_confidence') == 'low':
                cell.fill = _WARN
    for i, width in enumerate([16, 26, 14, 20, 12, 60], 1):
        boundary.column_dimensions[get_column_letter(i)].width = width

    if issues:
        ws = wb.create_sheet('异常记录')
        _header(ws, ['文档ID', '字段/阶段', '说明'])
        for rr, issue in enumerate(issues, 2):
            for c, value in enumerate([issue.get('document_id', ''), issue.get('field', ''),
                                       issue.get('message', '')], 1):
                cell = ws.cell(rr, c, str(value))
                cell.font = Font(size=9)
                cell.fill = _WARN
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_title(raw, used):
    title = ''.join('_' if ch in '[]:*?/\\' else ch for ch in str(raw))[:31] or '明细'
    base, n = title, 2
    while title in used:
        suffix = '-%d' % n
        title = base[:31 - len(suffix)] + suffix
        n += 1
    return title
